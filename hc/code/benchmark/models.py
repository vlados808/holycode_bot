import re
import csv
import datetime
from urllib.parse import urlparse
from hashlib import md5
from io import BytesIO
from pprint import pprint
import xml.etree.ElementTree as ET

from django.db import models
from django.forms.models import model_to_dict
from django.conf import settings
from django.db.utils import IntegrityError
from .parser import TelegramParser
from django.utils.html import format_html

import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup


class ParsingError(BaseException):
    pass


class TypeConversationError(BaseException):
    pass


class GradeConversationError(BaseException):
    pass


class VacancyType(models.Model):

    type = models.CharField(max_length=50, primary_key=True)
    manager = models.CharField(max_length=50, blank=True)
    kind = models.CharField(max_length=50)
    for_recruiter = models.BooleanField(null=True)
    name = models.CharField(max_length=50, blank=True)

    def __str__(self):
        return self.type


class Source(models.Model):

    class SourceType(models.TextChoices):
        TG = 'telegram'
        GS = 'google sheet'
        XL = 'file'
        XML = 'xml'

    name = models.CharField(max_length=50, primary_key=True)
    type = models.CharField(max_length=50, choices=SourceType.choices, blank=True)
    url = models.CharField(max_length=500, blank=True)
    instruction = models.JSONField()
    file = models.FileField(upload_to='vacancy_file', blank=True)

    @classmethod
    def create(cls, name):
        data = {
            'name': name,
            'type': 'telegram',
            'instruction': settings.DEFAULT_INSTRUCTION
        }
        obj = cls.objects.create(**data)
        print(f'NEW SOURCE ({obj}) CREATED')
        return obj

    @staticmethod
    def close_smartsourcing(data):
        reg = r'ID:? *(?P<source_number>[\d]+)'
        matches = re.finditer(reg, data, re.MULTILINE)
        for match in matches:
            value = match.group('source_number')
            print(f'VALUE {value}')
            vc = Vacancy.objects.filter(source='SmartSourcing', source_number=value).first()
            print(vc)
            vc.change_status('закрыта')

    @staticmethod
    def pause_smartsourcing(data):
        reg = r'ID:? *(?P<source_number>[\d]+)'
        matches = re.finditer(reg, data, re.MULTILINE)
        for match in matches:
            value = match.group('source_number')
            print(f'VALUE {value}')
            vc = Vacancy.objects.filter(source='SmartSourcing', source_number=value).first()
            print(vc)
            vc.change_status('приостановлена')

    @staticmethod
    def actualize_skillstaff(data):
        source = data.get('source')
        data = data.get('msg')
        all_db_vacancies = Vacancy.objects.filter(source=source)
        all_db_vacancies = {x.source_number: x for x in all_db_vacancies if x.source_number}

        reg = r'ID:? *(?P<source_number>[\d]+)'
        matches = re.finditer(reg, data, re.MULTILINE)
        for match in matches:
            value = match.group('source_number')
            print(f'VALUE {value}')
            try:
                all_db_vacancies.pop(value)
            except KeyError:
                text = f'{source} vacancy {value} not found'
                print(text)
                url = f'{settings.BOT_URL}send_absence'
                response = requests.post(url, json={"text": text})

        for _, vc in all_db_vacancies.items():
            if vc.status != 'закрыта':
                vc.change_status('закрыта')

    @staticmethod
    def actualize_ssp(data):
        source = data.get('source')
        data = data.get('msg')
        all_db_vacancies = Vacancy.objects.filter(source=source)
        all_db_vacancies = {x.source_number: x for x in all_db_vacancies if x.source_number}

        reg = r'^- (?P<source_number>[\d]+)'
        matches = re.finditer(reg, data, re.MULTILINE)
        for match in matches:
            value = match.group('source_number')
            print(f'VALUE {value}')
            try:
                all_db_vacancies.pop(value)
            except KeyError:
                text = f'{source} vacancy {value} not found'
                print(text)
                url = f'{settings.BOT_URL}send_absence'
                response = requests.post(url, json={"text": text})

        for _, vc in all_db_vacancies.items():
            if vc.status != 'закрыта':
                vc.change_status('закрыта')

    def actualize(self):
        all_db_vacancies = Vacancy.objects.filter(source=self)
        all_db_vacancies = {x.vacancy_id: x for x in all_db_vacancies}

        source_vacancies = self.get_source_vacancies()
        print(f'ALL DB - {len(all_db_vacancies)}, ALL SOURCE - {len(source_vacancies)}')

        for source_vc in source_vacancies:
            # print('ACTUALIZE')
            # pprint(source_vc)
            try:
                db_vc = all_db_vacancies.pop(source_vc.get('vacancy_id'))
            except KeyError:
                if source_vc.get('status'):
                    try:
                        Vacancy.create_new(source_vc)
                    except IntegrityError:
                        print(f'KEY {source_vc.get("vacancy_id")} ALREADY EXIST')
                        continue
            else:
                if source_vc.get('status') is None:
                    db_vc.change_status('приостановлена')
                else:
                    db_vc.change_status('открыта')

        for _, db_vc in all_db_vacancies.items():
            if db_vc.status != 'закрыта':
                db_vc.change_status('закрыта')

    def get_source_vacancies(self):
        if self.type == 'xml':
            return self.get_xml_vacancies()
        # elif self.type == 'file':
        #     return self.get_file_vacancies()
        return self.get_gs_vacancies()

    def get_xml_vacancies(self):
        r = requests.get(self.url)
        r.encoding = 'utf-8'
        root = ET.fromstring(r.text)
        res = []

        for element in root[0].findall('item'):
            guid = element.find('guid').text.split('/')[-1]

            title = element.find('title').text
            print(f'GUID {guid} {title}')
            title = self.fetch_vacancy_type(title)
            if not title:
                continue

            text = element.find('{http://purl.org/rss/1.0/modules/content/}encoded').text
            soup = BeautifulSoup(text, 'html.parser')
            text = '\n'.join(soup.strings)

            sub_res = {
                'source': self,
                'update': datetime.datetime.today(),
                'text': text,
                'source_number': guid,
                'type': title
            }

            for param in self.instruction:
                value = text

                if param.get('value') and param.get('name'):
                    sub_res[param.get('name')] = param.get('value')
                    continue

                if reg := param.get('reg'):

                    matches = re.finditer(reg, value, re.MULTILINE)
                    values = [x.group() for x in matches]
                    if values:
                        value = values[0]
                    else:
                        value = ''
                    # value = '\n'.join([x.group() for x in matches])

                if param_type := param.get('type'):
                    if param_type == 'type':
                        value = self.fetch_vacancy_type(value)
                    elif param_type == 'grade':
                        value = self.fetch_vacancy_grade(value)
                    elif param_type == 'int' and value:
                        value = int(value)

                if param.get('required'):
                    if not value:
                        raise ParsingError(f'param ({param}) value({value})')

                if value:
                    value = Source.remove_emoji(value)

                # print(f'YO {param} {value}')
                # value = value.lower()

                if name := param.get('name'):
                    if value:
                        if sub_res.get(name):
                            sub_res[name] += value
                        else:
                            sub_res[name] = value

            Vacancy.convert_type(sub_res)
            Vacancy.convert_grade(sub_res)
            Vacancy.check_experience(sub_res)
            Vacancy.check_currancy(sub_res)
            Vacancy.set_status(sub_res)
            Vacancy.create_name(sub_res)
            Vacancy.create_id(sub_res)

            res.append(sub_res)

        return res

    def get_gs_vacancies(self):
        ws = self.get_ws()
        res = []
        for i, row in enumerate(ws.rows):
            if i > 100:
                break  # TODO
            if self.type == 'file':
                if row[0].value in ('Активные', 'Направление', 'Заказчик'):
                    continue
                if row[0].fill.bgColor.value == '00000000':
                    break
                if row[0].fill.bgColor.value == 'FFFF9900':
                    continue
            try:
                print(f'TRY TO PROCESS {row}')
                res.append(self.get_xls_line(row))
            except ParsingError as e:
                print(f'NOT PROCESSED {e}')
                # print(row)
            except TypeConversationError as e:
                print(f'TYPE ERROR {e}')
                # print(row)
        return res

    def get_ws(self):
        if self.type == 'file':
            wb = load_workbook(self.file.path)
            return wb.active
        url_parsed = urlparse(str(self.url))
        key = url_parsed.path.split('/')[-2]
        response = requests.get(f'https://docs.google.com/spreadsheet/ccc?key={key}&output=xlsx')
        res = response.content
        wb = load_workbook(filename=BytesIO(res), data_only=True)
        return wb.active

    def parse(self):

        if self.type == 'file':
            wb = load_workbook(self.file.path)
            ws = wb.active
            for row in ws.rows:
                if row[0].value in ('Активные', 'Направление', 'Заказчик'):
                    continue
                if row[0].fill.bgColor.value == '00000000':
                    break
                if row[0].fill.bgColor.value == 'FFFF9900':
                    continue
                try:
                    self.process_xls_line(row)
                except ParsingError as e:
                    print(f'NOT PROCESSED {e}')
                    print(row)
                except TypeConversationError as e:
                    print(f'TYPE ERROR {e}')

        elif self.type == 'google sheet':
            url_parsed = urlparse(str(self.url))
            key = url_parsed.path.split('/')[-2]

            response = requests.get(f'https://docs.google.com/spreadsheet/ccc?key={key}&output=xlsx')
            res = response.content
            wb = load_workbook(filename=BytesIO(res), data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.rows):
                if i > 100:
                    break
                try:
                    print(f'TRY TO PROCESS {row}')
                    self.process_xls_line(row)
                except ParsingError as e:
                    print(f'NOT PROCESSED {e}')
                    print(row)
                except TypeConversationError as e:
                    print(f'TYPE ERROR {e}')
                    print(row)

            # response = requests.get(f'https://docs.google.com/spreadsheet/ccc?key={key}&output=csv')
            # res = response.content.decode('utf-8')
            # cr = csv.reader(res.splitlines())
            # raw_data = list(cr)
            #
            # for line in raw_data:
            #     try:
            #         self.process_line(line, raw_data)
            #     except ParsingError as e:
            #         print(f'NOT PROCESSED {e}')
            #         print(line)
            #     except TypeConversationError as e:
            #         print(f'TYPE ERROR {e}')

    def parse_xls_line(self, raw_line):
        instruction = self.instruction
        sub_res = {
            'source': self,
            'update': datetime.datetime.today(),
            'text': [],
        }

        for param in instruction:

            print(f'PARAM {param}')
            if param.get('value') and param.get('name'):
                sub_res[param.get('name')] = param.get('value')
                continue

            line = raw_line
            value = line[param.get('column')].value
            print(f'VALUE ({value})')

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')



            if value is None:
                continue
            sub_res['text'].append(str(value))
            value = str(value)
            value = value.lower()

            if reg := param.get('reg'):
                reg = re.compile(reg)
                value = reg.findall(value)
                if value:
                    value = value[0]
                else:
                    value = ''

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if param_type := param.get('type'):
                if param_type == 'type':
                    value = self.fetch_vacancy_type(value)
                elif param_type == 'grade':
                    value = self.fetch_vacancy_grade(value)
                elif param_type == 'int' and value:
                    value = int(value)

            if name := param.get('name'):
                if value:
                    if prefix := param.get('prefix'):
                        value = f'{prefix}{value}'
                    if sub_res.get(name):
                        sub_res[name] = ' '.join([sub_res[name], value])
                    else:
                        sub_res[name] = value

        sub_res['text'] = '\n'.join(sub_res['text'])

    def get_xls_line(self, raw_line):
        instruction = self.instruction
        sub_res = {
            'source': self,
            'update': datetime.datetime.now(),
            'text': [],
        }
        if self.type == 'file':
            sub_res['status'] = 'открыта'

        for param in instruction:

            print(f'PARAM {param}')
            if param.get('value') and param.get('name'):
                sub_res[param.get('name')] = param.get('value')
                continue

            line = raw_line
            value = line[param.get('column')].value
            print(f'VALUE ({value})')

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if value is None:
                continue

            sub_res['text'].append(str(value))
            value = str(value)
            value = value.lower()

            if reg := param.get('reg'):
                reg = re.compile(reg)
                value = reg.findall(value)
                if value:
                    value = value[0]
                else:
                    value = ''

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if param_type := param.get('type'):
                if param_type == 'type':
                    value = self.fetch_vacancy_type(value)
                elif param_type == 'grade':
                    value = self.fetch_vacancy_grade(value)
                elif param_type == 'int' and value:
                    value = int(value)

            if name := param.get('name'):
                if value:
                    if prefix := param.get('prefix'):
                        value = f'{prefix}{value}'
                    if name == 'project' and len(value) > 500:
                        value = value[:500]
                    if sub_res.get(name):
                        sub_res[name] = ' '.join([sub_res[name], value])
                    else:
                        sub_res[name] = value

        sub_res['text'] = '\n'.join(sub_res['text'])
        Vacancy.convert_type(sub_res)
        Vacancy.convert_grade(sub_res)
        Vacancy.check_experience(sub_res)
        Vacancy.check_currancy(sub_res)
        Vacancy.check_status(sub_res)
        Vacancy.create_name(sub_res)
        Vacancy.create_id(sub_res)
        return sub_res

    def process_xls_line(self, raw_line):
        instruction = self.instruction
        sub_res = {
            'source': self,
            'update': datetime.datetime.today(),
            'text': [],
        }

        for param in instruction:

            print(f'PARAM {param}')
            if param.get('value') and param.get('name'):
                sub_res[param.get('name')] = param.get('value')
                continue

            line = raw_line
            value = line[param.get('column')].value
            print(f'VALUE ({value})')

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')



            if value is None:
                continue
            sub_res['text'].append(str(value))
            value = str(value)
            value = value.lower()

            if reg := param.get('reg'):
                reg = re.compile(reg)
                value = reg.findall(value)
                if value:
                    value = value[0]
                else:
                    value = ''

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if param_type := param.get('type'):
                if param_type == 'type':
                    value = self.fetch_vacancy_type(value)
                elif param_type == 'grade':
                    value = self.fetch_vacancy_grade(value)
                elif param_type == 'int' and value:
                    value = int(value)

            if name := param.get('name'):
                if value:
                    if prefix := param.get('prefix'):
                        value = f'{prefix}{value}'
                    if sub_res.get(name):
                        sub_res[name] = ' '.join([sub_res[name], value])
                    else:
                        sub_res[name] = value

        sub_res['text'] = '\n'.join(sub_res['text'])
        Vacancy.create(sub_res)

    def process_line(self, raw_line, raw_data):
        print('PROCESS')
        print(raw_line)
        instruction = self.instruction
        sub_res = {
            'source': self,
            'update': datetime.datetime.today(),
            'text': '\n'.join(raw_line),
        }

        for param in instruction:

            if param.get('value') and param.get('name'):
                sub_res[param.get('name')] = param.get('value')
                continue

            if param.get('line') != 'inline':
                line = raw_data[param.get('line')]
            else:
                line = raw_line

            value = line[param.get('column')]
            value = value.lower()

            # print(f'YOYO {param} ({value})')

            if reg := param.get('reg'):
                reg = re.compile(reg)
                value = reg.findall(value)
                if value:
                    value = value[0]
                else:
                    value = ''

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if param_type := param.get('type'):
                if value:
                    if param_type == 'int':
                        value = int(value)

            if name := param.get('name'):
                if value:
                    if sub_res.get(name):
                        sub_res[name] = ' '.join([sub_res[name], value])
                    else:
                        sub_res[name] = value

        Vacancy.create(sub_res)

    def fetch_kesmaty(self, msg):
        reg = 'http.+'
        matches = re.finditer(reg, msg, re.MULTILINE)
        values = [x.group() for x in matches]
        if values:
            value = values[0]
        else:
            value = ''
        if not value:
            return msg

        r = requests.get(value)
        if r.status_code != 200:
            return msg

        r.encoding = 'utf-8'
        html_text = r.text
        soup = BeautifulSoup(html_text, 'html.parser')
        data = soup.find_all("article", {"class": "article__content text"})
        if not data:
            return msg

        res = []
        for element in data[0].strings:
            if element in ('Требования:', 'Основные обязанности:', 'Описание:') or element.startswith('Локация:'):
                res.append('\n')
            res.append(element)
            res.append('\n')
        data = ''.join(res)
        msg = '\n'.join((msg, data))
        return msg

    def fetch_stratosfera(self, msg):
        reg = 'http.+'
        matches = re.finditer(reg, msg, re.MULTILINE)
        values = [x.group() for x in matches]
        if values:
            value = values[0]
        else:
            value = ''
        if not value:
            return msg

        r = requests.get(value)
        if r.status_code != 200:
            return msg

        r.encoding = 'utf-8'
        html_text = r.text
        soup = BeautifulSoup(html_text, 'html.parser')
        data = soup.find_all("div", {"class": "article__container"})
        if not data:
            return msg

        res = []
        for element in data[0].strings:
            # if element in ('Требования:', 'Основные обязанности:', 'Описание:') or element.startswith('Локация:'):
            #     res.append('\n')
            res.append(element)
            res.append('\n')
        data = ''.join(res)
        msg = '\n'.join((msg, data))
        return msg

    def parse_tg(self, msg):
        # if self.name == 'Kesmaty':
        #     msg = self.fetch_kesmaty(msg)
        # elif self.name == 'Стратосфера':
        #     msg = self.fetch_stratosfera(msg)
            # print(msg)
            # return

        # parsed = TelegramParser.parse_msg(msg)
        # s = 1/0
        print('PROCESS')
        print(msg)
        instruction = self.instruction
        sub_res = {
            'source': self,
            'update': datetime.datetime.today(),
            'text': msg,
        }
        # value = msg.lower()

        for param in instruction:
            value = msg

            if param.get('value') and param.get('name'):
                sub_res[param.get('name')] = param.get('value')
                continue

            if reg := param.get('reg'):

                matches = re.finditer(reg, value, re.MULTILINE)
                values = [x.group() for x in matches]
                if values:
                    value = values[0]
                else:
                    value = ''
                # value = '\n'.join([x.group() for x in matches])

            if param_type := param.get('type'):
                if param_type == 'type':
                    value = self.fetch_vacancy_type(value)
                elif param_type == 'grade':
                    value = self.fetch_vacancy_grade(value)
                elif param_type == 'int' and value:
                    value = int(value)

            if param.get('required'):
                if not value:
                    raise ParsingError(f'param ({param}) value({value})')

            if value:
                value = Source.remove_emoji(value)

            if value and isinstance(value, str):
                value = value.strip()

            # print(f'YO {param} {value}')
            # value = value.lower()

            if name := param.get('name'):
                if value:
                    if sub_res.get(name):
                        sub_res[name] += value
                    else:
                        sub_res[name] = value

        Vacancy.create(sub_res)

    @staticmethod
    def remove_emoji(text):
        text = str(text)
        emoji_pattern = re.compile("["
                                   u"\U0001F600-\U0001F64F"  # emoticons
                                   u"\U0001F300-\U0001F5FF"  # symbols & pictographs
                                   u"\U0001F680-\U0001F6FF"  # transport & map symbols
                                   u"\U0001F1E0-\U0001F1FF"  # flags (iOS)
                                   "]+", flags=re.UNICODE)
        res = emoji_pattern.sub(r'', text)
        return res

    @staticmethod
    def fetch_vacancy_type(data):
        # print(f'FETCH TYPE {data}')
        data = data.lower()
        regs = [
            "системн[\\w]* аналитик[\\w]*|аналитик[\\w]* системн[\\w]*|systemanalyst|",
            "тестиров[\\w]* ручно[\\w]*|ручно[\\w]* тестиров[\\w]*|",
            "тестиров[\\w]* авто|авто тестиров[\\w]*|",
            "нагрузочного тестирования|",
            "тестирование|",
            "бизнес[ -]?аналитик[\\w]*|",
            "data аналитик|",
            "data инженер|",
            "аналитик|",
            "bi разработчик|",
            "разработка etl|",
            "java[ ]*script|js|",
            "java|",
            "ios|",
            "android|",
            "\\.net|",
            "react|",
            "python|",
            "golang|go(?=[\\n ])|",
            "c\\+\\+|с\\+\\+|",
            "c#|с#|",
            "front ?end|",
            "full[ -]?stack|",
            "dev[ ]?ops|",
            "qa функц|",
            "qa|",
            "angular|",
            "1c|1с|",
            "flutter|",
            "ruby|ruby on rails|ror|",
            "php|",
            "project manager|",
            "дизайнер|",
            "oracle",
        ]
        reg_s = "".join(regs)
        reg = re.compile(reg_s)
        value = reg.findall(data)
        print(f'TEST DATA {value}')
        if not value:
            return 'UNKNOWN'
            # print(reg_s)
            # print(data)
            # raise TypeConversationError("TYPE regex found 0 matches!")
        value = value[0]
        value = value.strip()
        return value

    @staticmethod
    def fetch_vacancy_grade(data):
        data = data.lower()
        regs = [
            "junior|джуниор|джун|",
            "mid{1,2}l?e? ?\\+?|",
            "мид{1,2}л? ?\\+?|",
            "senior|с[ие]н[иь][оеё]р",
        ]
        reg_s = "".join(regs)
        reg = re.compile(reg_s)
        value = reg.findall(data)
        if not value:
            return ''
            # print(reg_s)
            # print(data)
            # raise GradeConversationError("GRADE regex found 0 matches!")
        value = value[0]
        value = value.strip()
        return value

    def __str__(self):
        return self.name


class Status(models.TextChoices):
    OPEN = 'открыта'
    CLOSE = 'закрыта'
    CANCEL = 'отменена'
    IGNORE = 'не учитывать'
    PAUSE = 'приостановлена'


class Vacancy(models.Model):

    class Currencies(models.TextChoices):
        RUB = 'rubble'
        USD = 'dollar'

    source_number = models.CharField(max_length=50, blank=True)
    source = models.ForeignKey(Source, on_delete=models.CASCADE, null=True)
    contact = models.CharField(max_length=50, blank=True)
    vacancy_id = models.CharField(max_length=32, primary_key=True)
    short_id = models.CharField(max_length=8, null=True, blank=True)
    type = models.ForeignKey(VacancyType, on_delete=models.SET_NULL, null=True)
    status = models.CharField(max_length=50, choices=Status.choices, blank=True)
    priority = models.IntegerField(blank=True, default=0)
    created = models.DateTimeField(max_length=50, null=True)
    update = models.DateTimeField(max_length=50)
    last_publication = models.DateTimeField(max_length=50, null=True)

    name = models.CharField(max_length=50)
    grade = models.CharField(max_length=50)
    experience = models.CharField(max_length=50)
    rate = models.IntegerField(blank=True, default=0)
    partner_rate = models.IntegerField(blank=True, default=0)
    rate_currency = models.CharField(max_length=50, choices=Currencies.choices, blank=True)
    quantity = models.IntegerField(blank=True, default=1)

    project_term = models.CharField(max_length=500, blank=True)
    location = models.CharField(max_length=500, blank=True)
    project = models.CharField(max_length=500, blank=True)
    project_desc = models.TextField(blank=True)
    tasks = models.TextField(blank=True)
    requirements = models.TextField(blank=True)
    additional = models.TextField(blank=True)
    text = models.TextField(blank=True)

    class Meta:
        verbose_name_plural = 'Vacancies'

    def get_id_parts(self):
        if self.source.name in ('Потребности BSL', 'itMegabox Resources Requests'):
            return self.text
        salt = self.project_desc if self.project_desc else self.project
        value = ''.join([
            self.source.name,
            self.source_number,
            self.name,
            salt
        ])
        return value

    def manual_id(self):
        print('manual_id')
        value_to_encode = self.get_id_parts()
        value = value_to_encode.encode('utf-8')
        value = md5(value)
        value = value.hexdigest()
        return value
        # self.vacancy_id = value
        # self.save()

    def change_status(self, status):
        if self.status == status or self.status == 'не учитывать':
            return
        now = datetime.datetime.now()
        data = {
            'vacancy': self,
            'status': status,
            'date': now
        }
        VacancyStatus.objects.create(**data)
        self.update = now
        self.status = status
        self.save()

    @classmethod
    def create_new(cls, data):
        data['created'] = data['update']
        try:
            vc = cls.objects.create(**data)
        except Exception as e:
            print(f'EXCEPTION {e}')
            res = {x: len(y) for x, y in data.items() if type(y) == str}
            print(res)
            raise e
        print(f"vacancy {data['vacancy_id']} created")

        status_data = {
            'vacancy': vc,
            'status': 'открыта',
            'date': data.get('update')
        }
        VacancyStatus.objects.create(**status_data)

        vc.send('new')

    @classmethod
    def create(cls, data):
        print('PRERES')
        print(data)
        cls.convert_type(data)
        cls.convert_grade(data)
        cls.check_experience(data)
        cls.check_currancy(data)
        cls.set_status(data)
        cls.create_name(data)
        cls.create_id(data)

        try:
            vacancy_db = cls.objects.get(pk=data['vacancy_id'])
        except cls.DoesNotExist:
            data['created'] = data['update']
            vc_new = cls.objects.create(**data)
            print(f"vacancy {data['vacancy_id']} created")

            status_data = {
                'vacancy': vc_new,
                'status': 'открыта',
                'date': data.get('update')
            }
            VacancyStatus.objects.create(**status_data)

            vc_new.send('new')
            return

        # to_update = cls.find_changes(vacancy_db, data)
        # if to_update:
        #     cls.objects.filter(pk=data['vacancy_id']).update(**to_update)
        #     print(f"vacancy {data['vacancy_id']} updated with {to_update}")
        #     return

        print(f"vacancy {vacancy_db.vacancy_id} - {vacancy_db.short_id} is already exist")
        # cls.send_error(f'дублирование вакансии {vacancy_db.short_id}')

    @staticmethod
    def send_error(text):
        pass

    @staticmethod
    def convert_type(data):
        # print(data)
        type_map = {
            "JavaScript": "JS",
            "js": "JS",
            "javascript": "JS",
            "java script": "JS",
            "Тестировщик ручной": "Manual QA",
            "тестирование ручное": "Manual QA",
            "тестировщик авто": "Auto QA",
            "тестирование": "Manual QA",
            "нагрузочного тестирования": "Manual QA",
            "Системный аналитик": "System Analyst",
            "системный аналитик": "System Analyst",
            "системного аналитика": "System Analyst",
            "аналитик системный": "System Analyst",
            "аналитика системный": "System Analyst",
            "systemanalyst": "System Analyst",
            "data инженер": "Data",
            "data аналитик": "Data Analyst",
            "python": "Python",
            "bi разработчик": "DWH",
            "разработка etl": "DWH",
            "c++": "C++",
            "с++": "C++",
            "c#": ".NET",
            "с#": ".NET",
            "frontend": "FrontEnd",
            "front end": "FrontEnd",
            "full-stack": "FullStack",
            "fullstack": "FullStack",
            "flutter": "Flutter",
            "ios": "iOS",
            "разработчик ios": "iOS",
            "android": "Android",
            "разработчик android": "Android",
            "devops": "DevOps",
            "ror разработчик": "Ruby",
            "ror": "Ruby",
            "ruby": "Ruby",
            "ruby on rails": "Ruby",
            "java разработчик": "Java",
            "java": "Java",
            "разработчик java": "Java",
            ".net": ".NET",
            "qa manual": "Manual QA",
            "qa функц": "Manual QA",
            "qa": "Manual QA",
            "devops/sre": "DevOps",
            "ручной тестировщик": "Manual QA",
            "тестировщик ручной": "Manual QA",
            "react": "React",
            "разработчик react": "React",
            "angular": "Angular",
            "php": 'PHP',
            "golang": 'Go',
            "go": 'Go',
            "vue.js": 'Vue',
            "бизнес-аналитик": 'Business Analyst',
            "бизнес-аналитика": 'Business Analyst',
            "бизнес-аналитике": 'Business Analyst',
            "бизнес аналитик": 'Business Analyst',
            "бизнес аналитика": 'Business Analyst',
            "аналитик": 'Business Analyst',
            "1с": '1C',
            "1c": '1C',
            "project manager": 'Project Manager',
            "дизайнер": 'Design',
            "oracle": 'DevOps',
            "UNKNOWN": 'Unknown',
        }
        type_str = type_map.get(data.get('type'))
        if not type_str:
            raise TypeConversationError(f"Can't find type in mapping {data.get('type')}")
        try:
            type_db = VacancyType.objects.get(pk=type_str)
        except VacancyType.DoesNotExist:
            raise TypeConversationError(f"Can't find type in db {type_str}")
        data['type'] = type_db

    @staticmethod
    def convert_grade(data):
        grade_map = {
            'джун': 'Junior',
            'джуниор': 'Junior',
            'junior': 'Junior',
            'мид': 'Middle',
            'мидл': 'Middle',
            'миддл': 'Middle',
            'middle': 'Middle',
            'midle': 'Middle',
            'мид+': 'Middle+',
            'мид +': 'Middle+',
            'мидл+': 'Middle+',
            'мидл +': 'Middle+',
            'миддл+': 'Middle+',
            'миддл +': 'Middle+',
            'middle+': 'Middle+',
            'midle+': 'Middle+',
            'middle +': 'Middle+',
            'midle +': 'Middle+',
            'mid+': 'Middle+',
            'mid +': 'Middle+',
            'сениор': 'Senior',
            'сеньор': 'Senior',
            'сеньер': 'Senior',
            'сеньёр': 'Senior',
            'синиор': 'Senior',
            'синьор': 'Senior',
            'синьер': 'Senior',
            'синьёр': 'Senior',
            'senior': 'Senior',
            'архитектор': 'Architect',
            'architect': 'Architect',
        }
        grade_str = grade_map.get(data.get('grade'))
        if not grade_str:
            grade_str = 'Middle+'
            # raise GradeConversationError(f"Can't find grade in mapping {data.get('grade')}!")
        data['grade'] = grade_str

    @staticmethod
    def check_experience(data):
        exp_mapping = {
            'Junior': 'от 1 года',
            'Middle': 'от 2 лет',
            'Middle+': 'от 3 лет',
            'Senior': 'от 4 лет',
            'Architect': 'от 6 лет',
        }
        if data.get('experience'):
            return
        data['experience'] = exp_mapping.get(data.get('grade'), '')

    @staticmethod
    def create_name(data):
        name_parts = (x for x in (data.get('grade'), data['type'].type, data['type'].name) if x)
        data['name'] = ' '.join(name_parts)

    @staticmethod
    def create_id(data):
        val = data.get('text')

        # if data['source'].name in ('Потребности BSL', 'itMegabox Resources Requests'):
        #     val = data.get('text')
        # else:
        #     val = ''.join([
        #         data['source'].name,
        #         data.get('source_number', ''),
        #         data.get('name', ''),
        #         data.get('project_desc', data.get('project', '')),
        #     ])

        val = val.encode('utf-8')
        val = md5(val)
        val = val.hexdigest()
        data['vacancy_id'] = val

        now = datetime.datetime.now()
        data['short_id'] = f'HC{now.microsecond}'

    @staticmethod
    def find_changes(vacancy, data):
        vacancy_dict = model_to_dict(vacancy)
        for attr in (
                'source_number',
                'name',
                'project',
                'project_desc',
                'vacancy_id',
                'update',
                'priority',
                'status',
                'type',
                'source',
                'last_publication',
                'short_id',
        ):
            vacancy_dict.pop(attr)
        res = {}
        for key, value in vacancy_dict.items():
            if data.get(key, '') != value:
                if value == 0 and data.get(key) is None:
                    continue
                if key == 'quantity' and data.get(key) is None:
                    continue
                print(f"FIND CHANGES {key} {value} -> {data.get(key)}")
                res[key] = data.get(key, '')
        res['update'] = data.get('update')
        return res

    @staticmethod
    def set_status(data):
        data['status'] = 'открыта'

    @staticmethod
    def check_status(data):
        if data.get('status'):
            data['status'] = 'открыта'

    @staticmethod
    def check_currancy(data):
        if not data.get('rate_currency'):
            data['rate_currency'] = 'rubble'

    def send(self, status):
        endpoint_map = {
            'new': 'send_new',
            'pause': 'send_change',
            'close': 'send_change',
            'btn': 'vacancy',
        }
        url = f'{settings.BOT_URL}{endpoint_map.get(status)}'
        vacancy_dict = model_to_dict(self)
        vacancy_dict.pop('last_publication')
        vacancy_dict.pop('update')
        vacancy_dict.pop('created')
        vacancy_dict.pop('text')
        vacancy_dict['manager'] = self.type.manager

        response = requests.post(url, json=vacancy_dict)
        try:
            data = response.json()
        except Exception as e:
            print(f'EXCEPTION {status} -> {url} -> {response}')
            pprint(vacancy_dict)
            raise e
        print(f'BOT SEND {data}')

        pub_date = datetime.datetime.today()
        pub_data = {
            'last_published': pub_date,
            'vacancy': self,
            'chat_id': data.get('chat_id'),
            'message_id': data.get('message_id'),
            'partner_publication': True if status == 'btn' else False,
        }
        try:
            Publication.objects.create(**pub_data)
        except IntegrityError:
            pub_data.pop('vacancy')
            chat_id = pub_data.pop('chat_id')
            Publication.objects.filter(vacancy=self, chat_id=chat_id).update(**pub_data)

        self.last_publication = pub_date
        self.save()


class Publication(models.Model):
    vacancy = models.ForeignKey(Vacancy, on_delete=models.CASCADE, null=True)
    chat_id = models.BigIntegerField()
    last_published = models.DateTimeField()
    message_id = models.IntegerField()
    partner_publication = models.BooleanField(default=False)

    class Meta:
        unique_together = ('vacancy', 'chat_id',)


class VacancyStatus(models.Model):
    vacancy = models.ForeignKey(Vacancy, on_delete=models.CASCADE, null=True)
    status = models.CharField(max_length=50, choices=Status.choices, blank=True)
    date = models.DateTimeField()


class Partner(models.Model):
    nickname = models.CharField(max_length=50, primary_key=True)
    name = models.CharField(max_length=200)
    company_name = models.CharField(max_length=200)
    brand_name = models.CharField(max_length=200)
    registration_country = models.CharField(max_length=200)
    other_countries = models.CharField(max_length=200)
    inn = models.CharField(max_length=200)
    phone = models.CharField(max_length=200)
    email = models.CharField(max_length=200)


class Resume(models.Model):

    class ResumeStatus(models.TextChoices):
        New = 'New'
        InProgress = 'InProgress'
        Rejected = 'Rejected'
        Hold = 'Hold'

    vacancy = models.ForeignKey(Vacancy, on_delete=models.CASCADE, null=True)
    partner = models.ForeignKey(Partner, on_delete=models.CASCADE, null=True)
    date = models.DateTimeField(null=True)
    name = models.CharField(max_length=200)
    city = models.CharField(max_length=200)
    grade = models.CharField(max_length=200)
    rate = models.CharField(max_length=200)
    resume = models.FileField(upload_to='resume_dir', blank=True)
    status = models.CharField(max_length=50, choices=ResumeStatus.choices, default=ResumeStatus.New)
    additional = models.CharField(max_length=200, default='')

    def file_link(self):
        if self.resume:
            return format_html("<a href='%s'>download</a>" % (self.resume.url,))
        else:
            return "No attachment"

    file_link.allow_tags = True


class Digest(models.Model):
    date = models.DateTimeField()
